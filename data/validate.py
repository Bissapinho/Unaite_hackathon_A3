"""validate.py — vérifications de cohérence de la data synthétique (Brique 1).

Rapport PASS/FAIL sur 10 familles de checks. Code retour non-zéro si un check échoue.

Usage : python data/validate.py
Prérequis : avoir lancé `python data/generate_all.py` au moins une fois.
"""

from __future__ import annotations

import json
import pathlib
import sqlite3
import sys
from datetime import date, datetime, timedelta

ROOT = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from data import canonical as C  # noqa: E402

DATA = ROOT / "data"

# --------------------------------------------------------------------------- #
# Petit framework de check
# --------------------------------------------------------------------------- #
_RESULTS: list[tuple[bool, str, str]] = []


def check(cond: bool, label: str, detail: str = "") -> bool:
    _RESULTS.append((bool(cond), label, detail))
    return bool(cond)


def section(title: str) -> None:
    _RESULTS.append((None, f"__SECTION__{title}", ""))


# --------------------------------------------------------------------------- #
# Chargement des sources produites
# --------------------------------------------------------------------------- #
def load_json(path: pathlib.Path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_sources() -> dict:
    src = {}
    src["odoo"] = load_json(DATA / "odoo" / "odoo_dump.json")
    src["dashdoc"] = load_json(DATA / "dashdoc" / "dashdoc_dump.json")
    src["emails"] = load_json(DATA / "emails" / "emails.json")
    return src


# --------------------------------------------------------------------------- #
# 1. Recoupement inter-sources (mêmes IDs, vocabulaires différents)
# --------------------------------------------------------------------------- #
def check_cross_source(src: dict) -> None:
    section("1. Recoupement inter-sources (SH-2049 / MedPharma / PO-8821 / INV-7742)")

    odoo = src["odoo"]
    partners = {p["partner_id"]: p for p in odoo["res.partner"]}
    orders = {o["order_id"]: o for o in odoo["sale.order"]}
    moves = {m["move_id"]: m for m in odoo["account.move"]}
    products = {p["product_id"]: p for p in odoo["product.product"]}

    # ERP
    check("C001" in partners and partners["C001"]["name"] == "MedPharma",
          "ERP: MedPharma (C001) présent", str(partners.get("C001", {}).get("name")))
    check(orders.get("O-881", {}).get("client_order_ref") == "PO-8821",
          "ERP: O-881 -> PO-8821", str(orders.get("O-881", {}).get("client_order_ref")))
    check(moves.get("INV-7742", {}).get("amount_total") == 186000,
          "ERP: INV-7742 = 186000", str(moves.get("INV-7742", {}).get("amount_total")))
    check("PHARMA-22" in products,
          "ERP: PHARMA-22 présent", "")

    # TMS
    transports = {t["uid"]: t for t in src["dashdoc"]["transports"]}
    sh = transports.get("SH-2049", {})
    check("SH-2049" in transports, "TMS: SH-2049 présent", "")
    deliv = sh.get("deliveries", [])
    has_pharma = any(d.get("sku") == "PHARMA-22" and d.get("quantity") == 1000 for d in deliv)
    check(has_pharma, "TMS: SH-2049 contient PHARMA-22 x1000", str(deliv))
    check(sh.get("carrier", {}).get("name") == "ColdRoad",
          "TMS: SH-2049 carrier = ColdRoad", str(sh.get("carrier")))

    # Emails
    bodies = " ".join(e["subject"] + " " + e["body"] for e in src["emails"])
    check("SH-2049" in bodies, "Email: SH-2049 mentionné", "")
    check("PO-8821" in bodies, "Email: PO-8821 mentionné", "")

    # PDF (texte)
    pdf_text = extract_all_pdf_text()
    check("PO-8821" in pdf_text and "INV-7742" in pdf_text and "SH-2049" in pdf_text,
          "PDF: PO-8821 + INV-7742 + SH-2049 présents", "")

    # SQLite (legacy variante MedPharma)
    names = legacy_names()
    check(any("med" in n.lower() and "pharma" in n.lower() for n in names),
          "SQLite: référence MedPharma (variante) présente", "")


# --------------------------------------------------------------------------- #
# 2. Intégrité référentielle (scénario + bruit)
# --------------------------------------------------------------------------- #
def check_referential() -> None:
    section("2. Intégrité référentielle (canonical)")

    cust_ids = {c["customer_id"] for c in C.CUSTOMERS}
    skus = {p["sku"] for p in C.PRODUCTS}
    carrier_ids = {c["carrier_id"] for c in C.CARRIERS}
    wh_ids = {w["warehouse_id"] for w in C.WAREHOUSES}
    order_ids = {o["order_id"] for o in C.ORDERS}
    shipment_ids = {s["shipment_id"] for s in C.SHIPMENTS}

    check(all(o["customer_id"] in cust_ids for o in C.ORDERS),
          "Chaque order -> customer existe", "")
    check(all(s["order_id"] in order_ids for s in C.SHIPMENTS),
          "Chaque shipment -> order existe", "")
    check(all(s["carrier_id"] in carrier_ids for s in C.SHIPMENTS),
          "Chaque shipment -> carrier existe", "")
    check(all(s["origin_warehouse"] in wh_ids for s in C.SHIPMENTS),
          "Chaque shipment -> warehouse existe", "")
    check(all(it["sku"] in skus for it in C.SHIPMENT_ITEMS),
          "Chaque shipment_item -> sku existe", "")
    check(all(it["shipment_id"] in shipment_ids for it in C.SHIPMENT_ITEMS),
          "Chaque shipment_item -> shipment existe", "")
    check(all(inv["order_id"] in order_ids for inv in C.INVOICES),
          "Chaque invoice -> order existe", "")
    check(all(r["sku"] in skus and r["warehouse_id"] in wh_ids for r in C.INVENTORY),
          "Chaque inventory row -> sku + warehouse existent", "")
    check(all(ct["customer_id"] in cust_ids for ct in C.CONTRACTS),
          "Chaque contract -> customer existe", "")


# --------------------------------------------------------------------------- #
# 3. Saillance du scénario (SH-2049 ressort sans tricher)
# --------------------------------------------------------------------------- #
def check_salience() -> None:
    section("3. Saillance du scénario (SH-2049 = LE point chaud)")

    # MedPharma a la strategic_value max
    svals = [(c["customer_id"], c["strategic_value"]) for c in C.CUSTOMERS
             if c["strategic_value"] is not None]
    max_cust = max(svals, key=lambda x: x[1])
    check(max_cust[0] == "C001",
          "MedPharma a la strategic_value maximale", f"max={max_cust}")
    others = [v for cid, v in svals if cid != "C001"]
    check(all(v < 1200000 for v in others) and (not others or max(others) <= 600000),
          "Aucun autre client > ~600000", f"others_max={max(others) if others else None}")

    # INV-7742 = plus grosse facture Pending
    pending = [(i["invoice_id"], i["amount"]) for i in C.INVOICES if i["status"] == "Pending"]
    biggest = max(pending, key=lambda x: x[1])
    check(biggest[0] == "INV-7742" and biggest[1] == 186000,
          "INV-7742 = plus grosse facture Pending", f"biggest={biggest}")

    # SH-2049 = seul shipment cumulant cold-chain + Platinum + pénalité>=7000 + deadline du jour
    def order_of(sid):
        s = next(s for s in C.SHIPMENTS if s["shipment_id"] == sid)
        return C.ORDER_BY_ID[s["order_id"]]

    def penalty_of(cust_id):
        cts = [ct for ct in C.CONTRACTS if ct["customer_id"] == cust_id]
        return max((ct["late_penalty_per_hour"] for ct in cts), default=0)

    today_iso = C.TODAY.isoformat()
    cumul = []
    for s in C.SHIPMENTS:
        o = C.ORDER_BY_ID[s["order_id"]]
        cust = C.CUSTOMER_BY_ID[o["customer_id"]]
        is_cold = s["temperature_controlled"]
        is_platinum = cust["priority_tier"] == "Platinum"
        big_penalty = penalty_of(cust["customer_id"]) >= 7000
        today_deadline = o["promised_delivery_deadline"].startswith(today_iso)
        if is_cold and is_platinum and big_penalty and today_deadline:
            cumul.append(s["shipment_id"])
    check(cumul == ["SH-2049"],
          "SH-2049 seul à cumuler cold-chain+Platinum+pénalité>=7000+deadline du jour",
          f"cumul={cumul}")


# --------------------------------------------------------------------------- #
# 4. Cohérence interne du scénario
# --------------------------------------------------------------------------- #
def check_scenario_consistency() -> None:
    section("4. Cohérence scénario")

    sh2049 = next(s for s in C.SHIPMENTS if s["shipment_id"] == "SH-2049")
    check(sh2049["status"] == "Delayed", "SH-2049 status = Delayed", sh2049["status"])
    check(sh2049["order_id"] == "O-881", "SH-2049 -> O-881", sh2049["order_id"])

    o881 = C.ORDER_BY_ID["O-881"]
    check(o881["po_number"] == "PO-8821", "O-881 -> PO-8821", o881["po_number"])

    inv = next(i for i in C.INVOICES if i["order_id"] == "O-881")
    check(inv["invoice_id"] == "INV-7742" and inv["amount"] == 186000,
          "Invoice de O-881 = INV-7742 (186000)", f"{inv['invoice_id']}/{inv['amount']}")

    ct = next(c for c in C.CONTRACTS if c["customer_id"] == "C001")
    check(ct["late_penalty_per_hour"] == 7000 and ct["sla_deadline"] == "18:00",
          "Contrat C001 = 7000/h & deadline 18:00", f"{ct['late_penalty_per_hour']}/{ct['sla_deadline']}")

    invrow = next(r for r in C.INVENTORY if r["warehouse_id"] == "WH-2" and r["sku"] == "PHARMA-22")
    check(invrow["available_units"] == 200 and invrow["reserved_units"] == 100,
          "Inventory WH-2/PHARMA-22 = 200/100", f"{invrow['available_units']}/{invrow['reserved_units']}")


# --------------------------------------------------------------------------- #
# 5. Résolution floue (legacy_contacts)
# --------------------------------------------------------------------------- #
def legacy_names() -> list[str]:
    con = sqlite3.connect(DATA / "db" / "annex.db")
    try:
        rows = con.execute("SELECT raw_name FROM legacy_contacts").fetchall()
    finally:
        con.close()
    return [r[0] for r in rows]


def check_fuzzy() -> None:
    section("5. Résolution floue (legacy_contacts)")
    names = legacy_names()
    canonical_name = "MedPharma"
    variants = [n for n in names if "med" in n.lower() and "pharma" in n.lower()
                and n != canonical_name]
    check(len(variants) >= 1,
          "Variante orthographique de MedPharma présente et != entrée canonique",
          f"variants={variants}")


# --------------------------------------------------------------------------- #
# 6. PDF lisibles
# --------------------------------------------------------------------------- #
def extract_pdf_text(path: pathlib.Path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def extract_all_pdf_text() -> str:
    pdfs = sorted((DATA / "pdfs").glob("*.pdf"))
    return "\n".join(extract_pdf_text(p) for p in pdfs)


def _norm(t: str) -> str:
    # normalise espaces insécables / variantes d'espacement issues de l'extraction
    return t.replace(" ", " ").replace(" ", " ").replace(" ", "")


def check_pdfs() -> None:
    section("6. PDF lisibles (ré-extraction texte)")
    text = extract_all_pdf_text()
    n = _norm(text)
    tokens = ["PO-8821", "INV-7742", "€186,000", "SH-2049", "2°C", "8°C", "€7,000"]
    for tok in tokens:
        check(_norm(tok) in n, f"PDF contient « {tok} »", "")


# --------------------------------------------------------------------------- #
# 7. Excel lisibles
# --------------------------------------------------------------------------- #
def check_excel() -> None:
    section("7. Excel lisibles (en-têtes + lignes)")
    from openpyxl import load_workbook

    expected = {
        "carrier_backup_matrix.xlsx": (
            ["Route", "Backup Carrier", "Max Cold Chain Capacity", "Emergency Rate"],
            len(C.CARRIER_BACKUP_MATRIX),
        ),
        "customer_priority_list.xlsx": (
            ["Customer", "Priority Tier", "Account Manager", "Escalation Required"],
            len(C.CUSTOMERS),
        ),
        "warehouse_inventory_snapshot.xlsx": (
            ["Warehouse", "SKU", "Available Units", "Reserved Units"],
            len(C.INVENTORY),
        ),
    }
    for fname, (headers, nrows) in expected.items():
        wb = load_workbook(DATA / "excel" / fname, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        got_headers = list(rows[0]) if rows else []
        check(got_headers == headers, f"{fname}: en-têtes", f"{got_headers}")
        check(len(rows) - 1 == nrows, f"{fname}: {nrows} lignes", f"got={len(rows) - 1}")


# --------------------------------------------------------------------------- #
# 8. Volume cible (±20%)
# --------------------------------------------------------------------------- #
def check_volume() -> None:
    section("8. Volume cible (±20%)")
    targets = {
        "customers": 18, "products": 12, "orders": 28, "shipments": 28,
        "carriers": 10, "vehicles": 10, "drivers": 14, "warehouses": 4,
        "inventory": 25, "suppliers": 8, "contracts": 8, "invoices": 28,
        "claims": 12, "sla_penalties": 6,
    }
    got = C.counts()
    for key, target in targets.items():
        lo, hi = target * 0.8, target * 1.2
        v = got[key]
        check(lo <= v <= hi, f"Volume {key} ~ {target} (±20%)", f"got={v}")


# --------------------------------------------------------------------------- #
# 9. Dates ISO 8601 parsables (anti-régression du bug "heure 24" invalide)
# --------------------------------------------------------------------------- #
def _parse_iso(value: str) -> bool:
    """True si `value` est une date OU un datetime ISO 8601 parsable."""
    try:
        if "T" in value:
            datetime.fromisoformat(value)
        else:
            date.fromisoformat(value)
        return True
    except (ValueError, TypeError):
        return False


def check_iso_dates(src: dict) -> None:
    section("9. Dates ISO parsables")

    bad: list[str] = []

    def consider(source: str, field: str, value) -> None:
        if value is None:
            return
        if not isinstance(value, str):
            bad.append(f"{source}.{field}={value!r} (type {type(value).__name__})")
            return
        if not _parse_iso(value):
            bad.append(f"{source}.{field}={value!r}")

    # Dashdoc : tracking.eta de chaque transport
    for t in src["dashdoc"]["transports"]:
        consider(f"dashdoc.transport[{t.get('uid')}]", "tracking.eta",
                 (t.get("tracking") or {}).get("eta"))

    # Odoo : date_order des sale.order
    for o in src["odoo"]["sale.order"]:
        consider(f"odoo.sale.order[{o.get('order_id')}]", "date_order", o.get("date_order"))

    # Emails : received_at
    for e in src["emails"]:
        consider(f"email[{e.get('id')}]", "received_at", e.get("received_at"))

    # Canonical : deadlines / arrivals / order_date
    for s in C.SHIPMENTS:
        consider(f"canonical.shipment[{s['shipment_id']}]", "estimated_arrival",
                 s.get("estimated_arrival"))
    for o in C.ORDERS:
        consider(f"canonical.order[{o['order_id']}]", "promised_delivery_deadline",
                 o.get("promised_delivery_deadline"))
        consider(f"canonical.order[{o['order_id']}]", "order_date", o.get("order_date"))

    check(not bad, "Toutes les dates des sources sont ISO 8601 parsables",
          "; ".join(bad))

    # Anti-régression ciblé : SH-2049.estimated_arrival = 00:00 du lendemain de TODAY
    sh = next(s for s in C.SHIPMENTS if s["shipment_id"] == "SH-2049")
    eta = sh.get("estimated_arrival")
    parsable = isinstance(eta, str) and _parse_iso(eta)
    expected = (C.TODAY + timedelta(days=1))
    is_midnight_tomorrow = False
    if parsable and "T" in eta:
        dt = datetime.fromisoformat(eta)
        is_midnight_tomorrow = (dt.date() == expected and dt.hour == 0
                                and dt.minute == 0 and dt.second == 0)
    check(parsable and is_midnight_tomorrow,
          "SH-2049.estimated_arrival parsable ET = 00:00 du lendemain",
          f"eta={eta!r}, attendu={expected.isoformat()}T00:00:00")


# --------------------------------------------------------------------------- #
# 10. RH & finances (organigramme caché + reconstructibilité + finances dérivées)
# --------------------------------------------------------------------------- #
def _email_bodies() -> str:
    return " \n ".join(e["subject"] + " " + e["body"] for e in C.EMAILS)


def _xlsx_cells_text(fname: str) -> str:
    """Concatène toutes les cellules (en-têtes + valeurs) d'un .xlsx en texte."""
    from openpyxl import load_workbook
    wb = load_workbook(DATA / "excel" / fname, read_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(values_only=True):
        out.extend(str(c) for c in row if c is not None)
    wb.close()
    return " ".join(out)


def check_hr_finance() -> None:
    section("10. RH & finances")

    emps = C.EMPLOYEES
    by_id = {e["employee_id"]: e for e in emps}

    # --- 10.1 Organigramme caché cohérent ---------------------------------- #
    roots = [e for e in emps if e["manager_id"] is None]
    check(len(roots) == 1, "Exactement un employé sans manager (le dirigeant)",
          f"roots={[e['employee_id'] for e in roots]}")
    dg = roots[0] if roots else None
    check(all(e["manager_id"] in by_id for e in emps if e["manager_id"] is not None),
          "Chaque manager_id non-nul existe dans EMPLOYEES", "")
    check(dg is not None and C.COMPANY["ceo_employee_id"] == dg["employee_id"],
          "COMPANY.ceo_employee_id pointe sur le dirigeant",
          f"ceo={C.COMPANY['ceo_employee_id']}")

    # Acyclique : remonter chaque chaîne jusqu'à la racine sans boucler.
    def reaches_root(e) -> bool:
        seen = set()
        cur = e
        while cur is not None and cur["manager_id"] is not None:
            if cur["employee_id"] in seen:
                return False
            seen.add(cur["employee_id"])
            cur = by_id.get(cur["manager_id"])
            if cur is None:
                return False
        return cur is not None  # a atteint la racine
    check(all(reaches_root(e) for e in emps), "Organigramme acyclique (aucun cycle)", "")

    # title_rank cohérent : le manager a un rang strictement supérieur (numéro <).
    rank_ok = all(by_id[e["manager_id"]]["title_rank"] < e["title_rank"]
                  for e in emps if e["manager_id"] is not None)
    check(rank_ok, "title_rank cohérent (manager strictement supérieur)", "")

    # --- 10.2 Anti-fuite : l'organigramme n'est exposé NULLE PART ----------- #
    from openpyxl import load_workbook
    wb = load_workbook(DATA / "excel" / "company_directory.xlsx", read_only=True)
    hdr = [str(c) if c is not None else "" for c in next(wb.active.iter_rows(values_only=True))]
    wb.close()
    forbidden = ["manager", "n+1", "responsable hiérarchique", "reports", "reports_to",
                 "supérieur"]
    leak_hdr = [w for w in forbidden if any(w in h.lower() for h in hdr)]
    check(not leak_hdr, "company_directory.xlsx : aucune en-tête de hiérarchie",
          f"headers={hdr}, leaks={leak_hdr}")

    # Aucune clé `manager_id` ni reports_to dans un artefact sérialisé (xlsx + JSON).
    serialized = []
    for fx in (DATA / "excel").glob("*.xlsx"):
        serialized.append(_xlsx_cells_text(fx.name))
    for fj in DATA.rglob("*.json"):
        serialized.append(fj.read_text(encoding="utf-8"))
    blob = " ".join(serialized).lower()
    check("manager_id" not in blob and "reports_to" not in blob,
          "Aucune clé manager_id / reports_to dans les artefacts sérialisés (xlsx+json)",
          "")

    # --- 10.3 Reconstructibilité (heuristique de référence à deux branches) - #
    bodies = _email_bodies()
    direction = [e for e in emps if e["org_unit"] == "Direction"]
    by_unit: dict[str, list] = {}
    for e in emps:
        by_unit.setdefault(e["org_unit"], []).append(e)

    def is_chef(e) -> bool:
        if e["org_unit"] == "Direction":
            return False
        return e["title_rank"] == min(x["title_rank"] for x in by_unit[e["org_unit"]])

    def email_names_manager(e, mgr) -> bool:
        # un email d'escalade/délégation nomme le n+1 réel (full_name présent).
        return mgr["full_name"] in bodies

    unresolved = []
    for e in emps:
        if e["manager_id"] is None:
            continue
        mgr = by_id[e["manager_id"]]
        if is_chef(e):
            # (i) chef de service -> doit reporter à l'unique employé "Direction" (DG)
            ok = len(direction) == 1 and mgr["employee_id"] == direction[0]["employee_id"]
        else:
            # (ii) candidat = même org_unit, title_rank immédiatement supérieur
            sup_ranks = [x["title_rank"] for x in by_unit[e["org_unit"]]
                         if x["title_rank"] < e["title_rank"]]
            cand_rank = max(sup_ranks) if sup_ranks else None
            candidates = [x for x in by_unit[e["org_unit"]]
                          if x["title_rank"] == cand_rank]
            if len(candidates) == 1:
                ok = candidates[0]["employee_id"] == e["manager_id"]
            else:
                # (iii) ambiguïté -> exige un email nommant le bon n+1
                ok = email_names_manager(e, mgr)
        if not ok:
            unresolved.append(e["employee_id"])
    check(not unresolved,
          "Reconstructibilité : 100% des liens manager_id retrouvés (règle 2 branches)",
          f"non résolus={unresolved}")

    # --- 10.4 Indices emails (signatures + escalades reports_to) ------------ #
    signed = [e for e in emps
              if e["full_name"] in bodies and e["role_title"] in bodies]
    check(len(signed) >= len(emps) // 2,
          "Signatures « Nom — Poste » pour une majorité d'employés",
          f"signés={len(signed)}/{len(emps)}")

    def escalation_link(emp_name: str, mgr) -> bool:
        kw = ["valid", "transmet", "escalad", "remonte", "délég", "delegu",
              "prend", "arbitr", "accord"]
        for e in C.EMAILS:
            b = (e["subject"] + " " + e["body"]).lower()
            if emp_name.lower() in b and mgr["full_name"].lower() in b \
                    and any(k in b for k in kw):
                return True
        return False

    sarah = next(e for e in emps if e["full_name"] == "Sarah Martin")
    sarah_mgr = by_id[sarah["manager_id"]]
    check(escalation_link("Sarah Martin", sarah_mgr),
          "Email reliant Sarah Martin à son n+1 (manager réel nommé)",
          f"n+1={sarah_mgr['full_name']}")

    drv_emps = [e for e in emps if e["role_title"] == "Chauffeur"]
    drv_linked = any(escalation_link(d["full_name"], by_id[d["manager_id"]])
                     for d in drv_emps)
    check(drv_linked, "Email reliant un chauffeur fixé à son n+1 (manager réel nommé)", "")

    # --- 10.5 Réutilisation des account managers --------------------------- #
    ams = {c["account_manager"] for c in C.CUSTOMERS}
    emp_names = {e["full_name"] for e in emps}
    missing_am = sorted(ams - emp_names)
    check(not missing_am, "Chaque account_manager a un employé correspondant",
          f"manquants={missing_am}")

    # --- 10.6 Chauffeurs reliés (Driver is_a Employee) --------------------- #
    driver_names = {d["name"] for d in C.DRIVERS}
    linked_drivers = [e for e in drv_emps if e["full_name"] in driver_names]
    check(len(linked_drivers) >= 3,
          "≥ 3 chauffeurs EMPLOYEES dont le nom apparaît dans DRIVERS",
          f"reliés={len(linked_drivers)}")

    # --- 10.7 Cohérence masse salariale ------------------------------------ #
    payroll_sum = sum(e["monthly_gross_salary"] for e in emps)
    ratio = C.FINANCIAL_SUMMARY["payroll_monthly"] / payroll_sum
    check(1.0 <= ratio <= 1.6,
          "payroll_monthly ≈ somme des salaires (chargement patronal 1.0–1.6x)",
          f"ratio={ratio:.2f}")

    # --- 10.8 CA dérivé des factures --------------------------------------- #
    check(C.total_revenue() == sum(i["amount"] for i in C.INVOICES),
          "total_revenue() == somme INVOICES[].amount", "")
    conc = C.revenue_concentration()
    revs = [r["revenue"] for r in conc]
    check(revs == sorted(revs, reverse=True),
          "revenue_concentration() triée décroissante", "")
    # C001 (MedPharma) est un fait DÉRIVÉ des factures : présent dans le haut du
    # classement, sans forcer la 1re place (cf. DATA_README — strategic_value != CA).
    top_ids = [r["customer_id"] for r in conc[:5]]
    check("C001" in top_ids, "MedPharma (C001) dans le top 5 du CA (dérivé)",
          f"top5={top_ids}")

    # --- 10.9 Excel relisibles --------------------------------------------- #
    expected = {
        "company_directory.xlsx": (
            ["Matricule", "Nom", "Poste", "Service", "Email", "Date d'entrée"],
            len(C.EMPLOYEES),
        ),
    }
    for fname, (headers, nrows) in expected.items():
        wb = load_workbook(DATA / "excel" / fname, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        got_headers = list(rows[0]) if rows else []
        check(got_headers == headers, f"{fname}: en-têtes", f"{got_headers}")
        check(len(rows) - 1 == nrows, f"{fname}: {nrows} lignes", f"got={len(rows) - 1}")
    # finances_summary.xlsx s'ouvre et contient le CA dérivé + un bloc concentration
    fin_text = _xlsx_cells_text("finances_summary.xlsx")
    check("Concentration du CA par client" in fin_text and "MedPharma" in fin_text,
          "finances_summary.xlsx : bloc concentration présent (MedPharma listé)", "")

    # --- 10.10 Non-régression saillance (rien n'a bougé) ------------------- #
    svals = [(c["customer_id"], c["strategic_value"]) for c in C.CUSTOMERS
             if c["strategic_value"] is not None]
    check(max(svals, key=lambda x: x[1])[0] == "C001",
          "Non-régression : MedPharma garde la strategic_value max", "")
    pending = [(i["invoice_id"], i["amount"]) for i in C.INVOICES if i["status"] == "Pending"]
    check(max(pending, key=lambda x: x[1])[0] == "INV-7742",
          "Non-régression : INV-7742 reste la plus grosse facture Pending", "")
    sh = next(s for s in C.SHIPMENTS if s["shipment_id"] == "SH-2049")
    check(sh["status"] == "Delayed", "Non-régression : SH-2049 toujours Delayed", "")


# --------------------------------------------------------------------------- #
# Runner
# --------------------------------------------------------------------------- #
def main() -> int:
    src = load_sources()
    check_cross_source(src)
    check_referential()
    check_salience()
    check_scenario_consistency()
    check_fuzzy()
    check_pdfs()
    check_excel()
    check_volume()
    check_iso_dates(src)
    check_hr_finance()

    print("\n" + "=" * 72)
    print("RAPPORT DE VALIDATION — Brique 1 (data synthétique SH-2049)")
    print("=" * 72)
    n_pass = n_fail = 0
    for ok, label, detail in _RESULTS:
        if ok is None:
            print(f"\n{label.replace('__SECTION__', '## ')}")
            continue
        tag = "PASS" if ok else "FAIL"
        line = f"  [{tag}] {label}"
        if (not ok) and detail:
            line += f"  -> {detail}"
        print(line)
        if ok:
            n_pass += 1
        else:
            n_fail += 1

    print("\n" + "-" * 72)
    print(f"TOTAL: {n_pass} PASS / {n_fail} FAIL")
    print("=" * 72)
    return 0 if n_fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
